"""Convert every .xlsx / .xlsm file in an input folder to JSON in an output folder.

Each sheet of a workbook is written as its own JSON file named
`{workbook_stem}__{sheet_name}.json`, containing a list of row objects.
The first non-empty row of each sheet is treated as the header row.

Lifecycle of a source file when `processing_dir` and `processed_dir` are set:
  input_dir  -> (claim)   -> processing_dir
  processing_dir -> (ok)  -> processed_dir   (via finalize_to_processed_dir)
  processing_dir -> (err) -> stays in processing_dir for inspection

When `processing_dir` and `processed_dir` are both omitted, source files are
read in place and never moved.
"""

from __future__ import annotations

import datetime as dt
import json
import logging
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


logger = logging.getLogger(__name__)


class ExcelToJsonConverter:
    def __init__(
        self,
        input_dir: Path,
        output_dir: Path,
        processing_dir: Path | None = None,
        processed_dir: Path | None = None,
        clean_output_first: bool = False,
    ) -> None:
        if (processing_dir is None) != (processed_dir is None):
            raise ValueError(
                "processing_dir and processed_dir must be provided together, or both omitted"
            )
        self.input_dir = input_dir
        self.output_dir = output_dir
        self.processing_dir = processing_dir
        self.processed_dir = processed_dir
        self.clean_output_first = clean_output_first
        # Files successfully converted during this object's lifetime, sitting
        # in processing_dir, waiting for finalize_to_processed_dir() to move
        # them to processed_dir. Failed conversions stay in processing_dir but
        # are NOT listed here.
        self.successfully_processed_paths: list[Path] = []

    @property
    def _moves_sources(self) -> bool:
        return self.processing_dir is not None and self.processed_dir is not None

    def wipe_output(self) -> int:
        """Delete every top-level file from output_dir. Subdirectories are
        left alone. Returns the number of files deleted. Used by callers that
        need per-file isolation of output_dir (call before each process_one).
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)
        wiped = 0
        for p in self.output_dir.iterdir():
            if p.is_file():
                p.unlink()
                wiped += 1
        logger.info("wiped %d file(s) from %s", wiped, self.output_dir)
        return wiped

    def process_one(self, xlsx: Path) -> bool:
        """Process a single xlsx end-to-end: claim source to processing_dir
        (when move-mode), read the workbook, write per-sheet JSONs into
        output_dir, and record the source in successfully_processed_paths for
        later finalize_to_processed_dir().

        Does NOT wipe output_dir — the caller is responsible for that. Use
        wipe_output() before each process_one() call if you need output_dir
        to contain only this one file's sheets at a time.

        Returns True on success, False on conversion failure (already logged).
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)
        if self._moves_sources:
            self.processing_dir.mkdir(parents=True, exist_ok=True)
            self.processed_dir.mkdir(parents=True, exist_ok=True)
            source = self.processing_dir / xlsx.name
            # When processing_dir == input_dir, source == xlsx and the rename
            # would be a no-op (or platform-dependent error). Skip it.
            # shutil.move handles cross-volume falls back to copy+unlink.
            if source != xlsx:
                shutil.move(str(xlsx), str(source))
        else:
            source = xlsx

        try:
            data = self.convert_workbook(source)
        except Exception as exc:
            suffix = f" (left in {source})" if self._moves_sources else ""
            logger.error("FAILED %s: %s%s", xlsx.name, exc, suffix)
            return False

        written: list[str] = []
        for sheet_name, records in data.items():
            target = self.output_dir / f"{xlsx.stem}__{self.safe_name(sheet_name)}.json"
            with target.open("w", encoding="utf-8") as f:
                json.dump(records, f, ensure_ascii=False, indent=2)
            written.append(target.name)

        files_msg = ", ".join(written) or "(no sheets)"
        if self._moves_sources:
            self.successfully_processed_paths.append(source)
            logger.info(
                "%s -> %s (held in %s, pending finalize)",
                xlsx.name, files_msg, source.parent.name,
            )
        else:
            logger.info("%s -> %s", xlsx.name, files_msg)
        return True

    def convert_folder(self) -> int:
        """Batch-process every .xls[xm] in input_dir. Wipes output_dir once
        upfront (if clean_output_first), so output_dir ends up containing the
        UNION of all successfully processed files' sheet JSONs.

        For per-file isolation (output_dir contains exactly one input file's
        sheets at a time), call wipe_output() + process_one() in your own
        loop instead — see run.py's primary stage.
        """
        if not self.input_dir.is_dir():
            logger.error("Input directory not found: %s", self.input_dir)
            return 1

        if self.clean_output_first:
            self.wipe_output()

        source_files = sorted(
            p for p in self.input_dir.glob("*.xls[xm]") if not p.name.startswith("~$")
        )
        if not source_files:
            logger.info("No .xlsx/.xlsm files found in %s", self.input_dir)
            return 0

        for xlsx in source_files:
            self.process_one(xlsx)

        return 0

    def unmark_processed(self, name: str) -> int:
        """Remove every entry from `successfully_processed_paths` whose
        basename matches `name`. Returns the count removed.

        Use when a downstream stage (e.g. TIA generation) failed for this
        file so it should NOT graduate to processed_dir at finalize time —
        leaving it in processing_dir lets the operator retry on the next
        scheduled run. No-op if the converter wasn't configured to move
        sources.
        """
        if not self._moves_sources:
            return 0
        before = len(self.successfully_processed_paths)
        self.successfully_processed_paths = [
            p for p in self.successfully_processed_paths if p.name != name
        ]
        return before - len(self.successfully_processed_paths)

    def finalize_to_processed_dir(self) -> int:
        """Move every file recorded in `successfully_processed_paths` from
        processing_dir to processed_dir. Returns the number of files moved.

        Intended to be called once at the end of the overall pipeline run, so
        source files only graduate to processed_dir after the whole run has
        had a chance to use them. No-op if the converter wasn't configured
        with processing_dir / processed_dir, or if nothing was successfully
        processed.
        """
        if not self._moves_sources:
            return 0

        moved = 0
        for src in list(self.successfully_processed_paths):
            if not src.is_file():
                logger.warning(
                    "finalize skipped: %s no longer at %s", src.name, src
                )
                continue
            target = self.processed_dir / src.name
            try:
                # shutil.move handles cross-volume fallback to copy+unlink.
                # Delete a stale target first so the fallback won't refuse
                # to overwrite.
                if target.exists():
                    target.unlink()
                shutil.move(str(src), str(target))
            except Exception as exc:
                logger.error("finalize FAILED for %s: %s", src.name, exc)
                continue
            logger.info("finalized: %s -> %s", src.name, target)
            moved += 1

        # Clear after attempting the moves so a second finalize call doesn't
        # try to re-move (or warn about) the same paths.
        self.successfully_processed_paths = []
        return moved

    @staticmethod
    def safe_name(name: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_")
        return cleaned or "sheet"

    def convert_workbook(self, path: Path) -> dict[str, list[dict[str, Any]]]:
        """Read an .xlsx/.xlsm at `path` and return
        `{sheet_name: [row_dict, ...]}`. Each row dict drops cells whose
        value is None or "". Date/time/datetime cells are ISO-serialized.
        """
        wb = load_workbook(filename=path, data_only=True, read_only=True)
        try:
            return {name: self._sheet_to_records(wb[name]) for name in wb.sheetnames}
        finally:
            wb.close()

    def _sheet_to_records(self, sheet) -> list[dict[str, Any]]:
        headers: list[str] | None = None
        records: list[dict[str, Any]] = []

        for row in sheet.iter_rows(values_only=False):
            values = [self._cell_value(c) for c in row]
            if headers is None:
                if all(v is None or v == "" for v in values):
                    continue
                headers = [
                    str(v) if v is not None else f"column_{i + 1}"
                    for i, v in enumerate(values)
                ]
                continue

            if all(v is None or v == "" for v in values):
                continue

            record: dict[str, Any] = {}
            for i, header in enumerate(headers):
                value = values[i] if i < len(values) else None
                if value is None or value == "":
                    continue
                record[header] = value
            records.append(record)

        return records

    @staticmethod
    def _cell_value(cell: Cell) -> Any:
        value = cell.value
        if isinstance(value, (dt.datetime, dt.date, dt.time)):
            return value.isoformat()
        return value
