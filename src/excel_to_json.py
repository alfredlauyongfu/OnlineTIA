"""Stage customer input files from an input folder into JSON in an output folder.

Two input formats are supported (see CUSTOMER_INPUT_PATTERNS):
  - Excel workbooks (.xlsx/.xlsm): each sheet is written as its own JSON file
    named `{workbook_stem}__{sheet_name}.json`, containing a list of row
    objects. The first non-empty row of each sheet is treated as the header.
  - JSON form exports (.json): validated (must parse as a JSON object;
    a UTF-8 BOM is tolerated) and re-serialized into the output folder
    under the same basename.

Lifecycle of a source file when `processing_dir` and `processed_dir` are set:
  input_dir  -> (claim)   -> processing_dir
  processing_dir -> (ok)  -> processed_dir   (via finalize_to_processed_dir)
  processing_dir -> (err) -> stays in processing_dir for inspection

When `processing_dir` and `processed_dir` are both omitted, source files are
read in place and never moved.
"""

from __future__ import annotations

import datetime as dt
import json
import logging
import re
import shutil
import stat
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


logger = logging.getLogger(__name__)


# Customer inputs may be Excel workbooks (converted per-sheet) or JSON form
# exports (validated + staged as-is). The reference stage stays xlsx-only —
# see `list_workbooks`.
CUSTOMER_INPUT_PATTERNS: tuple[str, ...] = ("*.xls[xm]", "*.json")


def move_replacing(src: Path, dst: Path) -> None:
    """Move `src` to `dst`, replacing any stale `dst`. shutil.move handles
    cross-volume moves (e.g. local C: to a mapped network drive) by falling
    back to copy+unlink where Path.replace() would raise OSError; deleting a
    stale target first keeps that fallback from refusing to overwrite."""
    if dst.exists():
        dst.unlink()
    shutil.move(str(src), str(dst))


# --- OneDrive / Files On-Demand robustness --------------------------------
# A customer input synced via OneDrive can arrive as an "online-only"
# placeholder: its metadata is on disk (so claiming it into processing_dir is a
# fast rename that succeeds) but its bytes are not. The first content read then
# triggers a hydration/recall that fails with OSError [Errno 22] Invalid
# argument when no OneDrive client is available to service it — e.g. under the
# non-interactive service account that runs the scheduled pipeline. These
# Windows attributes mark such a placeholder; os.stat() reads them WITHOUT
# itself triggering a recall (so the check is safe and cheap). The attributes
# are absent on non-Windows, where the bitmask is harmless and the check is
# always False.
_PLACEHOLDER_ATTRS = (
    getattr(stat, "FILE_ATTRIBUTE_OFFLINE", 0x1000)
    | getattr(stat, "FILE_ATTRIBUTE_RECALL_ON_OPEN", 0x40000)
    | getattr(stat, "FILE_ATTRIBUTE_RECALL_ON_DATA_ACCESS", 0x400000)
)
_PLACEHOLDER_HINT = (
    "{err}: '{path}' is a OneDrive online-only placeholder — its content is not "
    "downloaded locally, so it cannot be read. In OneDrive, right-click the "
    "input folder and choose 'Always keep on this device' (or disable Files "
    "On-Demand), and make sure the OneDrive client is running and signed in for "
    "the account that runs this pipeline."
)
# A read is retried a few times to ride out a file OneDrive is still syncing
# down (transient lock / partial write); a detected placeholder fails fast with
# the actionable hint above, because retrying can never hydrate it.
_READ_RETRIES = 3
_READ_RETRY_DELAY = 2.0


def _is_cloud_placeholder(path: Path) -> bool:
    """True if `path` is a OneDrive / Files On-Demand online-only placeholder
    (Windows only; always False elsewhere). Reads metadata only — never
    triggers a recall."""
    try:
        attrs = getattr(path.stat(), "st_file_attributes", 0)
    except OSError:
        return False
    return bool(attrs & _PLACEHOLDER_ATTRS)


def read_json_object(path: Path) -> dict[str, Any]:
    """Read and JSON-parse `path` (tolerating a UTF-8 BOM) and return the
    object.

    Retries on OSError to ride out a file OneDrive is still syncing down. A
    detected online-only placeholder raises an actionable OSError immediately
    (instead of the opaque [Errno 22]) since retrying cannot help it. Raises
    ValueError if the payload is valid JSON but not an object.
    """
    for attempt in range(_READ_RETRIES):
        try:
            with path.open("r", encoding="utf-8-sig") as f:
                payload = json.load(f)
        except OSError as exc:
            if _is_cloud_placeholder(path):
                raise OSError(
                    _PLACEHOLDER_HINT.format(err=exc.strerror or exc, path=path)
                ) from exc
            if attempt + 1 >= _READ_RETRIES:
                raise
            logger.warning(
                "read of %s failed (%s); retry %d/%d in %.0fs",
                path.name, exc, attempt + 1, _READ_RETRIES - 1, _READ_RETRY_DELAY,
            )
            time.sleep(_READ_RETRY_DELAY)
            continue
        if not isinstance(payload, dict):
            raise ValueError(f"expected a JSON object, got {type(payload).__name__}")
        return payload
    raise AssertionError("unreachable: loop returns or raises on every path")


class ExcelToJsonConverter:
    def __init__(
        self,
        input_dir: Path,
        output_dir: Path,
        processing_dir: Path | None = None,
        processed_dir: Path | None = None,
        clean_output_first: bool = False,
    ) -> None:
        if (processing_dir is None) != (processed_dir is None):
            raise ValueError(
                "processing_dir and processed_dir must be provided together, or both omitted"
            )
        self.input_dir = input_dir
        self.output_dir = output_dir
        self.processing_dir = processing_dir
        self.processed_dir = processed_dir
        self.clean_output_first = clean_output_first
        # Files successfully converted during this object's lifetime, sitting
        # in processing_dir, waiting for finalize_to_processed_dir() to move
        # them to processed_dir. Failed conversions stay in processing_dir but
        # are NOT listed here.
        self.successfully_processed_paths: list[Path] = []

    @property
    def _moves_sources(self) -> bool:
        return self.processing_dir is not None and self.processed_dir is not None

    def wipe_output(self) -> int:
        """Delete every top-level file from output_dir. Subdirectories are
        left alone. Returns the number of files deleted. Used by callers that
        need per-file isolation of output_dir (call before each process_one).
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)
        wiped = 0
        for p in self.output_dir.iterdir():
            if p.is_file():
                p.unlink()
                wiped += 1
        logger.info("wiped %d file(s) from %s", wiped, self.output_dir)
        return wiped

    def process_one(self, source: Path) -> bool:
        """Stage a single customer input end-to-end: claim it to
        processing_dir (when move-mode), then either convert an Excel
        workbook to per-sheet JSONs or validate-and-stage a JSON form export
        into output_dir, recording the source in successfully_processed_paths
        for later finalize_to_processed_dir().

        Does NOT wipe output_dir — the caller is responsible for that. Use
        wipe_output() before each process_one() call if you need output_dir
        to contain only this one file's artifacts at a time.

        Returns True on success, False on staging failure (already logged).
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)
        if self._moves_sources:
            self.processing_dir.mkdir(parents=True, exist_ok=True)
            self.processed_dir.mkdir(parents=True, exist_ok=True)
            claimed = self.processing_dir / source.name
            # When processing_dir == input_dir, claimed == source and the
            # move would be a no-op (or platform-dependent error). Skip it.
            if claimed != source:
                move_replacing(source, claimed)
        else:
            claimed = source

        if claimed.suffix.lower() == ".json":
            written = self._stage_json(claimed)
        else:
            written = self._convert_excel(claimed)
        if written is None:
            return False

        files_msg = ", ".join(written) or "(no sheets)"
        if self._moves_sources:
            self.successfully_processed_paths.append(claimed)
            logger.info(
                "%s -> %s (held in %s, pending finalize)",
                source.name, files_msg, claimed.parent.name,
            )
        else:
            logger.info("%s -> %s", source.name, files_msg)
        return True

    def _convert_excel(self, source: Path) -> list[str] | None:
        """Excel branch of process_one: workbook → one JSON per sheet in
        output_dir. Returns the written filenames, or None on failure."""
        try:
            data = self.convert_workbook(source)
        except Exception as exc:
            suffix = f" (left in {source})" if self._moves_sources else ""
            logger.error("FAILED %s: %s%s", source.name, exc, suffix)
            return None

        written: list[str] = []
        for sheet_name, records in data.items():
            target = self.output_dir / f"{source.stem}__{self.safe_name(sheet_name)}.json"
            with target.open("w", encoding="utf-8") as f:
                json.dump(records, f, ensure_ascii=False, indent=2)
            written.append(target.name)
        return written

    def _stage_json(self, source: Path) -> list[str] | None:
        """JSON branch of process_one: validate the form export (must parse
        as a JSON object; utf-8-sig tolerates a BOM) and re-serialize it into
        output_dir under the same basename. Re-serializing — not copying —
        guarantees the staged file is plain UTF-8 and exactly what was
        validated, which is what the TIA generator's reader expects.
        Returns the written filename, or None on failure."""
        try:
            payload = read_json_object(source)
        except Exception as exc:
            suffix = f" (left in {source})" if self._moves_sources else ""
            logger.error("FAILED %s: %s%s", source.name, exc, suffix)
            return None

        target = self.output_dir / source.name
        with target.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return [target.name]

    def convert_folder(self) -> int:
        """Batch-process every .xls[xm] in input_dir. Wipes output_dir once
        upfront (if clean_output_first), so output_dir ends up containing the
        UNION of all successfully processed files' sheet JSONs.

        For per-file isolation (output_dir contains exactly one input file's
        sheets at a time), call wipe_output() + process_one() in your own
        loop instead — see run.py's primary stage.
        """
        if not self.input_dir.is_dir():
            logger.error("Input directory not found: %s", self.input_dir)
            return 1

        if self.clean_output_first:
            self.wipe_output()

        source_files = self.list_workbooks(self.input_dir)
        if not source_files:
            logger.info("No .xlsx/.xlsm files found in %s", self.input_dir)
            return 0

        for xlsx in source_files:
            self.process_one(xlsx)

        return 0

    def unmark_processed(self, name: str) -> int:
        """Remove every entry from `successfully_processed_paths` whose
        basename matches `name`. Returns the count removed.

        Use when a downstream stage (e.g. TIA generation) failed for this
        file so it should NOT graduate to processed_dir at finalize time —
        leaving it in processing_dir lets the operator retry on the next
        scheduled run. No-op if the converter wasn't configured to move
        sources.
        """
        if not self._moves_sources:
            return 0
        before = len(self.successfully_processed_paths)
        self.successfully_processed_paths = [
            p for p in self.successfully_processed_paths if p.name != name
        ]
        return before - len(self.successfully_processed_paths)

    def finalize_to_processed_dir(self) -> int:
        """Move every file recorded in `successfully_processed_paths` from
        processing_dir to processed_dir. Returns the number of files moved.

        Intended to be called once at the end of the overall pipeline run, so
        source files only graduate to processed_dir after the whole run has
        had a chance to use them. No-op if the converter wasn't configured
        with processing_dir / processed_dir, or if nothing was successfully
        processed.
        """
        if not self._moves_sources:
            return 0

        moved = 0
        for src in list(self.successfully_processed_paths):
            if not src.is_file():
                logger.warning(
                    "finalize skipped: %s no longer at %s", src.name, src
                )
                continue
            target = self.processed_dir / src.name
            try:
                move_replacing(src, target)
            except Exception as exc:
                logger.error("finalize FAILED for %s: %s", src.name, exc)
                continue
            logger.info("finalized: %s -> %s", src.name, target)
            moved += 1

        # Clear after attempting the moves so a second finalize call doesn't
        # try to re-move (or warn about) the same paths.
        self.successfully_processed_paths = []
        return moved

    @staticmethod
    def safe_name(name: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_")
        return cleaned or "sheet"

    @staticmethod
    def list_workbooks(dir_path: Path) -> list[Path]:
        """Every .xlsx/.xlsm in `dir_path` (sorted), excluding Office `~$`
        lock/temp files. A missing directory yields [] — callers keep their
        own policy for that case."""
        return sorted(
            p for p in dir_path.glob("*.xls[xm]") if not p.name.startswith("~$")
        )

    @staticmethod
    def list_customer_inputs(dir_path: Path) -> list[Path]:
        """Every customer input file in `dir_path` (sorted) matching any of
        CUSTOMER_INPUT_PATTERNS, excluding Office `~$` lock/temp files."""
        return sorted(
            p
            for pattern in CUSTOMER_INPUT_PATTERNS
            for p in dir_path.glob(pattern)
            if not p.name.startswith("~$")
        )

    @staticmethod
    def sheet_name_from_path(path: Path) -> str:
        """Inverse of the `{workbook_stem}__{sheet_name}.json` naming this
        converter writes: the sheet part after the FIRST `__`, or the whole
        stem when the separator is absent."""
        stem = path.stem
        return stem.split("__", 1)[1] if "__" in stem else stem

    def convert_workbook(self, path: Path) -> dict[str, list[dict[str, Any]]]:
        """Read an .xlsx/.xlsm at `path` and return
        `{sheet_name: [row_dict, ...]}`. Each row dict drops cells whose
        value is None or "". Date/time/datetime cells are ISO-serialized.
        """
        wb = load_workbook(filename=path, data_only=True, read_only=True)
        try:
            return {name: self._sheet_to_records(wb[name]) for name in wb.sheetnames}
        finally:
            wb.close()

    def _sheet_to_records(self, sheet) -> list[dict[str, Any]]:
        headers: list[str] | None = None
        records: list[dict[str, Any]] = []

        for row in sheet.iter_rows(values_only=True):
            values = [self._json_value(v) for v in row]
            if headers is None:
                if all(v is None or v == "" for v in values):
                    continue
                headers = [
                    str(v) if v is not None else f"column_{i + 1}"
                    for i, v in enumerate(values)
                ]
                continue

            if all(v is None or v == "" for v in values):
                continue

            record: dict[str, Any] = {}
            for i, header in enumerate(headers):
                value = values[i] if i < len(values) else None
                if value is None or value == "":
                    continue
                record[header] = value
            records.append(record)

        return records

    @staticmethod
    def _json_value(value: Any) -> Any:
        """Cell value → JSON-serializable value (date/time ISO-serialized)."""
        if isinstance(value, (dt.datetime, dt.date, dt.time)):
            return value.isoformat()
        return value
